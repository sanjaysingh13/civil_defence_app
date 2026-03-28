"""
Tests for the is_functional field on Equipment and the seed_equipment command.

Covers three areas:
  1. Model / factory behaviour — is_functional default, independence from status.
  2. Command helper functions — clean_equip_name, safe_int, parse_equipment_sheet.
  3. End-to-end command behaviour — unique_id format, functional assignment logic,
     new-unit creation, Kolkata skip, and idempotency.

The end-to-end tests build a minimal in-memory Excel file using openpyxl that
mirrors the real sheet's pivot structure, so we test the full parsing and seeding
code path without depending on the real xlsx file being present.
"""

from __future__ import annotations

import io

import openpyxl
import pytest
from django.core.management import call_command

from civil_defence_app.equipment.models import Equipment, EquipmentStatus
from civil_defence_app.incidents.tests.factories import EquipmentFactory, UnitFactory

pytestmark = pytest.mark.django_db


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _make_test_xlsx(rows: list[dict]) -> io.BytesIO:
    """
    Build a minimal in-memory xlsx file that matches the pivot structure that
    parse_equipment_sheet expects.

    Parameters
    ----------
    rows : list[dict]
        Each dict describes one district row with keys:
            district (str)     — goes in col 0
            circ_saw_total     — Circular Saw total count
            circ_saw_func      — Circular Saw functional count
            fire_axe_total     — Fire Axe total count
            fire_axe_func      — Fire Axe functional count

    The generated sheet layout:
        Row 1 (idx 0) : District | Circular Saw | <NaN> | Fire Axe | <NaN>
        Row 2 (idx 1) : <NaN>    | Total | Functional  | Total | Functional
        Row 3+ (idx 2+): data rows (one per district)

    We use equipment names that exist in EQUIP_META ("Circular Saw", "Fire Axe")
    so the command recognises them without triggering "unknown equipment" warnings.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # The command reads from the sheet named "All Item current" — set it explicitly.
    ws.title = "All Item current"

    # Row 0 — equipment category headers (equipment name in col B, D, ...)
    ws.cell(row=1, column=1, value="District")
    ws.cell(row=1, column=2, value="Circular Saw")
    ws.cell(row=1, column=3, value=None)
    ws.cell(row=1, column=4, value="Fire Axe")
    ws.cell(row=1, column=5, value=None)

    # Row 1 — Total / Functional sub-headers
    ws.cell(row=2, column=2, value="Total")
    ws.cell(row=2, column=3, value="Functional")
    ws.cell(row=2, column=4, value="Total")
    ws.cell(row=2, column=5, value="Functional")

    # Data rows
    for i, row in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=row["district"])
        ws.cell(row=i, column=2, value=row["circ_saw_total"])
        ws.cell(row=i, column=3, value=row["circ_saw_func"])
        ws.cell(row=i, column=4, value=row["fire_axe_total"])
        ws.cell(row=i, column=5, value=row["fire_axe_func"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_tmp_xlsx(tmp_path, rows: list[dict]):
    """
    Save the in-memory xlsx to a real file and return its Path.
    call_command can't accept a BytesIO object, so we need a real path.
    """
    buf = _make_test_xlsx(rows)
    p = tmp_path / "test_equipment.xlsx"
    p.write_bytes(buf.read())
    return p


# ─────────────────────────────────────────────────────────────────────────────
# 1. MODEL — is_functional field
# ─────────────────────────────────────────────────────────────────────────────

class TestIsFunctionalField:
    """Unit tests for the Equipment.is_functional BooleanField."""

    def test_factory_default_is_true(self):
        """
        EquipmentFactory must produce items with is_functional=True by default.
        This mirrors the status="OK" default and represents a working item.
        """
        equip = EquipmentFactory.create()
        assert equip.is_functional is True

    def test_can_create_non_functional(self):
        """
        An equipment item with is_functional=False must be saveable and
        retrievable from the database without constraint violations.
        """
        equip = EquipmentFactory.create(is_functional=False, status="REPAIR")
        reloaded = Equipment.objects.get(pk=equip.pk)
        assert reloaded.is_functional is False

    def test_is_functional_is_independent_of_status(self):
        """
        is_functional and status are separate DB columns.  They are usually kept
        in sync by the seeding command, but nothing at the model level enforces
        that they match — an item can have status=REPAIR but is_functional=True
        (e.g. repaired but not yet marked OK) without any integrity error.
        """
        equip = EquipmentFactory.create(is_functional=True, status="REPAIR")
        reloaded = Equipment.objects.get(pk=equip.pk)
        assert reloaded.is_functional is True
        assert reloaded.status == EquipmentStatus.REPAIR

    def test_database_default_true(self):
        """
        Equipment created via ORM directly (not the factory) must also default
        to is_functional=True thanks to the BooleanField(default=True) declaration.
        """
        unit = UnitFactory.create()
        equip = Equipment.objects.create(
            unit=unit,
            name="Test Rope",
            unique_id="TESTUNIT-ROPE-001",
        )
        assert equip.is_functional is True


# ─────────────────────────────────────────────────────────────────────────────
# 2. COMMAND HELPERS — clean_equip_name, safe_int
# ─────────────────────────────────────────────────────────────────────────────

class TestCommandHelpers:
    """
    Unit tests for the pure helper functions inside seed_equipment.py.
    We import them directly from the module so they can be tested in isolation,
    without invoking the full Django management command machinery.
    """

    def setup_method(self):
        # Import helpers lazily here so the test module itself doesn't fail if
        # the command file has a syntax error during collection.
        from civil_defence_app.equipment.management.commands.seed_equipment import (
            clean_equip_name,
            safe_int,
        )
        self.clean = clean_equip_name
        self.sint  = safe_int

    def test_clean_equip_name_strips_whitespace(self):
        """Leading and trailing spaces must be removed."""
        assert self.clean("  Circular Saw  ") == "Circular Saw"

    def test_clean_equip_name_collapses_internal_spaces(self):
        """
        The Excel headers sometimes have double spaces inside the name
        (e.g. 'Portable  Generator Set ').  clean_equip_name must collapse
        any run of whitespace to a single space so the EQUIP_META lookup works.
        """
        assert self.clean("Portable  Generator  Set ") == "Portable Generator Set"

    def test_clean_equip_name_handles_leading_space(self):
        """A single leading space (as seen in ' Gri-Gri') must be stripped."""
        assert self.clean(" Gri-Gri") == "Gri-Gri"

    def test_safe_int_normal_int(self):
        """Integer input passes through as-is."""
        assert self.sint(5) == 5

    def test_safe_int_float(self):
        """pandas reads numeric cells as float; 3.0 must become 3."""
        assert self.sint(3.0) == 3

    def test_safe_int_nan_returns_zero(self):
        """NaN (empty cell) must become 0, not raise."""
        import math
        assert self.sint(math.nan) == 0

    def test_safe_int_negative_clamps_to_zero(self):
        """Negative values are nonsensical for counts; they must be clamped."""
        assert self.sint(-2) == 0

    def test_safe_int_string_returns_zero(self):
        """Non-numeric strings (e.g. 'N/A') must return 0."""
        assert self.sint("N/A") == 0


# ─────────────────────────────────────────────────────────────────────────────
# 3. COMMAND — end-to-end behaviour
# ─────────────────────────────────────────────────────────────────────────────

class TestSeedEquipmentCommand:
    """
    End-to-end tests for the seed_equipment management command.

    Each test builds a small temp xlsx file using _write_tmp_xlsx and then
    calls the command via call_command().  The unit name "Alipurduar" is used
    because it exists in UNIT_NAME_MAP and maps to the DB name "ALIPURDUAR".
    """

    def test_dry_run_creates_no_records(self, tmp_path):
        """
        --dry-run must parse and print stats but must NOT write any
        Equipment or Unit rows to the database.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 3, "circ_saw_func": 2,
             "fire_axe_total": 1, "fire_axe_func": 1},
        ])
        before = Equipment.objects.count()
        call_command("seed_equipment", xlsx=str(xlsx), dry_run=True)
        assert Equipment.objects.count() == before

    def test_creates_correct_number_of_records(self, tmp_path):
        """
        Total number of Equipment rows created must equal the sum of all
        'total' counts across all (district × equipment type) pairs with qty > 0.
        Here: 3 Circular Saws + 1 Fire Axe = 4 rows for Alipurduar.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 3, "circ_saw_func": 2,
             "fire_axe_total": 1, "fire_axe_func": 1},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        assert Equipment.objects.count() == 4

    def test_unique_id_format(self, tmp_path):
        """
        unique_id must follow the pattern UNIT_SLUG_UPPER-EQUIP_CODE-SL_NO,
        e.g. 'ALIPURDUAR-CIRC-SAW-001' for the first Circular Saw of Alipurduar.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 1, "circ_saw_func": 1,
             "fire_axe_total": 0, "fire_axe_func": 0},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        assert Equipment.objects.filter(unique_id="ALIPURDUAR-CIRC-SAW-001").exists()

    def test_functional_assignment_older_items_non_functional(self, tmp_path):
        """
        Given total=5, functional=3 the FIRST (5-3)=2 items must be
        is_functional=False (older, presumed worn out) and the last 3 must be
        is_functional=True (newer, presumed still working).
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 5, "circ_saw_func": 3,
             "fire_axe_total": 0, "fire_axe_func": 0},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))

        # SL 001 and 002 → non-functional (older)
        for sl in ["001", "002"]:
            eq = Equipment.objects.get(unique_id=f"ALIPURDUAR-CIRC-SAW-{sl}")
            assert eq.is_functional is False, f"SL {sl} should be non-functional"
            assert eq.status == EquipmentStatus.REPAIR

        # SL 003, 004, 005 → functional (newer)
        for sl in ["003", "004", "005"]:
            eq = Equipment.objects.get(unique_id=f"ALIPURDUAR-CIRC-SAW-{sl}")
            assert eq.is_functional is True, f"SL {sl} should be functional"
            assert eq.status == EquipmentStatus.FUNCTIONAL

    def test_all_functional_when_total_equals_functional(self, tmp_path):
        """
        When total == functional all items must be is_functional=True
        (no non-functional count, so nothing ages out to False).
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 3, "circ_saw_func": 3,
             "fire_axe_total": 0, "fire_axe_func": 0},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        for sl in ["001", "002", "003"]:
            eq = Equipment.objects.get(unique_id=f"ALIPURDUAR-CIRC-SAW-{sl}")
            assert eq.is_functional is True

    def test_none_functional_when_functional_is_zero(self, tmp_path):
        """
        When functional=0 all items must be is_functional=False
        (total items exist but none work).
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 2, "circ_saw_func": 0,
             "fire_axe_total": 0, "fire_axe_func": 0},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        for sl in ["001", "002"]:
            eq = Equipment.objects.get(unique_id=f"ALIPURDUAR-CIRC-SAW-{sl}")
            assert eq.is_functional is False

    def test_new_unit_is_created(self, tmp_path):
        """
        A district like 'Purulia' (not seeded from volunteer data) must be
        created as a new Unit row when the command runs for the first time.
        """
        from civil_defence_app.personnel.models import Unit

        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Purulia", "circ_saw_total": 2, "circ_saw_func": 1,
             "fire_axe_total": 0, "fire_axe_func": 0},
        ])
        assert not Unit.objects.filter(name="PURULIA").exists()
        call_command("seed_equipment", xlsx=str(xlsx))
        assert Unit.objects.filter(name="PURULIA").exists()

    def test_kolkata_is_skipped(self, tmp_path):
        """
        'Kolkata' maps to None in UNIT_NAME_MAP because it covers three DB
        units (CSA, NSA, SSA) that cannot be automatically split.  The command
        must skip it without error and create zero equipment rows for it.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Kolkata", "circ_saw_total": 5, "circ_saw_func": 5,
             "fire_axe_total": 3, "fire_axe_func": 3},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        # No Equipment rows should be created (Kolkata skipped)
        assert Equipment.objects.count() == 0

    def test_total_row_is_skipped(self, tmp_path):
        """
        The bottom 'Total' summary row in the sheet must be ignored — it is
        a grand sum, not a real unit.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 2, "circ_saw_func": 2,
             "fire_axe_total": 0, "fire_axe_func": 0},
            {"district": "Total",     "circ_saw_total": 2, "circ_saw_func": 2,
             "fire_axe_total": 0, "fire_axe_func": 0},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        # Only Alipurduar's 2 saws — Total row must add nothing
        assert Equipment.objects.count() == 2

    def test_idempotent_rerun_does_not_duplicate(self, tmp_path):
        """
        Running seed_equipment twice on the same file must produce the same
        set of Equipment rows — update_or_create on unique_id prevents duplication.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 3, "circ_saw_func": 2,
             "fire_axe_total": 0, "fire_axe_func": 0},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        first_run_count = Equipment.objects.count()
        call_command("seed_equipment", xlsx=str(xlsx))
        assert Equipment.objects.count() == first_run_count

    def test_category_assigned_correctly(self, tmp_path):
        """
        Equipment rows must be tagged with the category from EQUIP_META.
        Circular Saw → RESCUE, Fire Axe → FIRE.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 1, "circ_saw_func": 1,
             "fire_axe_total": 1, "fire_axe_func": 1},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        saw = Equipment.objects.get(unique_id="ALIPURDUAR-CIRC-SAW-001")
        axe = Equipment.objects.get(unique_id="ALIPURDUAR-FIRE-AXE-001")
        assert saw.category == "RESCUE"
        assert axe.category == "FIRE"

    def test_zero_total_rows_are_skipped(self, tmp_path):
        """
        Cells where total=0 produce no Equipment rows — there is nothing to import.
        """
        xlsx = _write_tmp_xlsx(tmp_path, [
            {"district": "Alipurduar", "circ_saw_total": 0, "circ_saw_func": 0,
             "fire_axe_total": 2, "fire_axe_func": 1},
        ])
        call_command("seed_equipment", xlsx=str(xlsx))
        # Only 2 Fire Axe rows — no Circular Saw rows at all
        assert Equipment.objects.filter(name__icontains="Circular").count() == 0
        assert Equipment.objects.filter(name__icontains="Fire Axe").count() == 2
